from main.ems_main import *
from main.ems_cover import *
from main.ems_chapt_central import chapt_search
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font

term = Terminal()

class CustomTheme(Default):
    def __init__(self):
        super().__init__() 
        self.List.selection_cursor = "➤"
        self.Checkbox.selection_icon = "➤"
        self.List.selection_color = term.green
        self.Checkbox.selection_color = term.green
        self.Checkbox.selected_color = term.green
        self.Checkbox.unselected_icon = "*"

def name_code(name):
    f = pd.read_excel("origin.xlsx", sheet_name="SETTINGS",usecols="A,O")
    df = dict(zip(f.edit_name, f.Manga))
    return df[name]

def excel_col_name(x):
    if x==0:
        return()
    if 1<=x<=26:
        return chr(64+x)
    r = int(x%26)
    k = int((x-r)/26)
    if r==0:
        r=26
        k-=1
        return excel_col_name(k)+'Z'
    else:
        return excel_col_name(k)+excel_col_name(r)

# ========================== #

def manga_select():

    f = pd.read_excel("origin.xlsx", sheet_name="SETTINGS",usecols="O,N,H")
    df_tbd = f[f['up.py']==True][['edit_name','TBD']]
    df_tbd.set_index('edit_name',drop=True,inplace=True)
    df = f[f['up.py']==True][['edit_name']]
    df.sort_values('edit_name',inplace=True)
    manga_list = df['edit_name'].to_list()
    q_manga = [
    inquirer.List(name='manga_choice',
                    message="Quel manga à DL",
                    choices=manga_list+['↪ Quit'],
                    carousel=True,
                )
    ]

    a_manga = inquirer.prompt(q_manga, theme=CustomTheme(), raise_keyboard_interrupt=True)['manga_choice']
    if a_manga == "↪ Quit":
        quit()
    tbd_exist = df_tbd.loc[a_manga]['TBD']
    if tbd_exist == True:
        scan_mode_ch = ['all','TBD update','select']
    else:
        scan_mode_ch = ['all','select']
    q_param = [
    inquirer.List(name='cover_update',
                     message="Update covers",
                     choices=['Yes','No'],
                     carousel=True),
    inquirer.List(name='scan_mode',
                     message="Scan update mode",
                     choices=scan_mode_ch,
                     carousel=True)
    ]
    p_param = inquirer.prompt(q_param, theme=CustomTheme(), raise_keyboard_interrupt=True)
    a_cover, a_mode = p_param['cover_update'],p_param['scan_mode']
    a_manga = name_code(a_manga)
    a_cover = (a_cover=='Yes')
    if a_mode == 'select':
        q_scan = [
            inquirer.Text("start_scan", message="Start scan"),
            inquirer.Text("end_scan", message="End scan [int ou 'max']")
        ]
        p_scan = inquirer.prompt(q_scan, theme=CustomTheme(), raise_keyboard_interrupt=True)
        if p_scan['end_scan'].isnumeric() == False:
            a_scan = [int(p_scan['start_scan']),'max']
        else:
            a_scan = [int(p_scan['start_scan']),int(p_scan['end_scan'])]
        return (a_manga, a_cover, a_scan)
    else:
        return(a_manga,a_cover, a_mode)

def edit_origin():
    
    def value_update(a:str):
        out = []
        for x in a.split(','):
            try:
                out.append(int(x))
            except:
                if x.lower() in ["na","n/a","n#a","#na"]:
                    print('ok')
                    out.append("=NA()")
                else:
                    raise ValueError
        return out

    f = pd.read_excel("origin.xlsx", sheet_name="SETTINGS",usecols="O,C")
    df = f[f["Chapt"].isin(["F"])==False][['edit_name']]
    df.sort_values('edit_name',inplace=True)
    manga_list = df['edit_name'].to_list()

    q_manga = [
    inquirer.List(name='manga_choice',
                    message="Manga à update dans origin",
                    choices=manga_list+['↪ Quit'],
                    carousel=True,
                )
    ]
    a_manga = inquirer.prompt(q_manga, theme=CustomTheme(), raise_keyboard_interrupt=True)['manga_choice']
    if a_manga == "↪ Quit":
        quit()
    a_manga = name_code(a_manga)

    wb = openpyxl.load_workbook('origin.xlsx')
    ws = wb["UPDATE"]
    ml = [c.value for c in ws["1"] if c.value != None]
    col_n = ml.index(a_manga)+1

    g = pd.read_excel("origin.xlsx", sheet_name="UPDATE",usecols="A,"+excel_col_name(col_n))

    pd_wb = g[['Vol']].copy()
    pd_wb["xl_rows"] = [k for k in range(2,len(pd_wb)+2)]
    pd_wb = pd_wb.set_index("Vol",drop=True)
    pd_wb = pd_wb.to_dict()["xl_rows"]

    g = g.set_index("Vol",drop=True)
    g.dropna(inplace=True)
    g = g.astype({a_manga: int})
    print(g.head().to_markdown())
    print('\n')

    decision = [
        inquirer.List('decision',choices=['Edit','↪ Quit'],carousel=True)
    ]
    dec = inquirer.prompt(decision, theme=CustomTheme(), raise_keyboard_interrupt=True)['decision']

    if dec == '↪ Quit':
        quit()

    q_update = [
        inquirer.Text("vols", message="Vol à maj 'x,x' "),
        inquirer.Text("chapts", message="Chapt fin 'x,x' ")
    ]
    a_update = inquirer.prompt(q_update, theme=CustomTheme(), raise_keyboard_interrupt=True)
    #----------
    vols = [x.replace(' ','') for x in a_update["vols"].split(',')]
    if 'q' in a_update["vols"] or 'q' in a_update["chapts"]:
        quit()
    for vol in vols:
        try:
            vols[vols.index(vol)] = int(vol)
        except:
            if vol.lower() != 'tbd':
                print(term.red('>> Value error in Volumes'))
                quit()
            else:
                vols[vols.index(vol)] = "TBD"
    
    #try:
    #    chapts = [int(x) for x in a_update["chapts"].split(',')]
    try:
        chapts = value_update(a_update["chapts"])
    except ValueError:
        print(term.red('>> Value error in Chapts'))
        quit()
    updates = dict(zip(vols,chapts))

    for v in updates.keys():
        ws.cell(row=pd_wb[v],column=col_n).value = updates[v]
        ws.cell(row=pd_wb[v],column=col_n).alignment = Alignment(horizontal="center", vertical="center")
    wb.save('origin.xlsx')
    g2 = pd.read_excel("origin.xlsx", sheet_name="UPDATE",usecols="A,"+excel_col_name(col_n))
    g2 = g2.set_index("Vol",drop=True)
    g2.dropna(inplace=True)
    g2 = g2.astype({a_manga: int})
    print(g2.head().to_markdown())
    print('\n')
        
    return()

def add_origin():
    q_manga = [
    inquirer.Text(name='manga_choice',message="Code manga à ajouter dans origin")
    ]
    a_manga = inquirer.prompt(q_manga, theme=CustomTheme(), raise_keyboard_interrupt=True)['manga_choice']

    wb = openpyxl.load_workbook('origin.xlsx')
    ws = wb["UPDATE"]
    ml = [c.value for c in ws["1"] if c.value != None]

    if (a_manga.lower() in [x.lower() for x in ml]) == True:
        print(term.steelblue3('➜ {} déjà dans UPDATE'.format(a_manga)))
        edit_origin()
    
    else:
        # +/- comme edit_origin()
        q_add = [
            inquirer.Text("vol_max", message="Volume max "),
            inquirer.Text("chapts", message="Chapts fin pour tous les tomes 'x,x' ")
        ]
        a_add = inquirer.prompt(q_add, theme=CustomTheme(), raise_keyboard_interrupt=True)

        vols = [k for k in range(1,int(a_add['vol_max'])+1)]
        try:
            chapts = [int(x) for x in a_add["chapts"].split(',')]
        except ValueError:
            print(term.red('>> Value error in Chapts'))
            quit()
        add = dict(zip(vols,chapts))
        pd_wb = pd.read_excel("origin.xlsx", sheet_name="UPDATE",usecols="A")
        pd_wb["xl_rows"] = [k for k in range(2,len(pd_wb)+2)]
        pd_wb = pd_wb.set_index("Vol",drop=True)
        pd_wb = pd_wb.to_dict()["xl_rows"]
        col_n = len(ml)+1

        nm = ws.cell(row=1,column=col_n)
        nm.value = a_manga
        nm.fill = PatternFill(fill_type='solid',start_color='00FFFF00',end_color='00FFFF00')
        nm.font = Font(bold=True)
        nm.alignment = Alignment(horizontal="center", vertical="center")

        other_rows = [k for k in range(2,len(pd_wb)+2)]
        for v in add.keys():
            ws.cell(row=pd_wb[v],column=col_n).value = add[v]
            ws.cell(row=pd_wb[v],column=col_n).alignment = Alignment(horizontal="center", vertical="center")
            other_rows.remove(pd_wb[v])
        for row in other_rows:
            ws.cell(row=row,column=col_n).value = "=NA()"
            if row == 2:
                ws.cell(row=row,column=col_n).fill = PatternFill(fill_type='solid',start_color='00FFFF00',end_color='00FFFF00')
        
        wb.save('origin.xlsx')
        #print(term.steelblue3("➜ Aller maj onglet SETTINGS"))
        
        # TODO : ajouter un tableau qui répertorie les dates d'update
        return()

def source_dl():

    f = pd.read_excel("origin.xlsx", sheet_name="SETTINGS",usecols="O,E,N,B")
    f = f[f["up.py"]==True]
    f.drop(labels="up.py",axis=1,inplace=True)
    f.sort_values('edit_name',inplace=True)
    m_list = f['edit_name'].to_list()

    q_source = [
        inquirer.Checkbox(name='source',
                      message='',
                      choices=m_list+['↪ Quit'])
    ]
    a_source = inquirer.prompt(q_source, theme=CustomTheme(), raise_keyboard_interrupt=True)["source"]

    if '↪ Quit' in a_source:
        quit()
    else:
        mask = f['edit_name'].isin(a_source)
        df = f[mask][['Manga_path','Source']].copy()
        df.sort_values('Source',inplace=True)
        df.set_index('Source',drop=True,inplace=True)
        print(df.to_markdown())
        print('\n')
    return()

def check_converted():  #* (pq?)
    dispo = []
    for el in os.listdir(converted):
        if ('Complete' in el) == False:
            dispo.append(el.split(' ',1)[1])
        else:
            dispo.append(el.split('Complete - ')[1])
    dispo.sort()
    return dispo

def tbd_replace(manga,pth):
    try:
        fold = [x for x in os.listdir(converted) if manga in x][0]
        try:os.remove(converted+fold+"/"+manga+" VolTBD.cbz")
        except:pass
        try:os.remove(converted+fold+"/"+manga+" VolTBD.zip")
        except:pass
        move(pth+'/'+manga+' VolTBD.zip',converted+fold+"/"+manga+" VolTBD.zip")
        rmtree(pth,ignore_errors=True)
    except:
        print("<Erreur> pas de dossier pour {0} dans $ '{1}'".format(manga,converted))

def check_updates(mode, check_view="short"):
    
    g = pd.read_excel("origin.xlsx", sheet_name="SETTINGS",usecols="A,B,F,Q,R")
    d_statut = g[["Manga_path","Statut"]].copy().set_index("Manga_path",drop=True).to_dict()["Statut"]
    chapt_origin = g[["Manga_path","chapt_origin"]].copy().set_index("Manga_path",drop=True).to_dict()["chapt_origin"]
    inpt = [x for x in os.listdir(base_path) if x[:2]!="._"]
    clean = [x[:len(x)-1] for x in os.listdir(clean_path) if x[:2]!="._"]
    s3 = set(g[g["update_monitor"]==False]["Manga_path"])
    s1,s2 = set(inpt), set(clean)
    
    def statut_check(x):
        nonlocal d_statut
        try:
            return d_statut[x]
        except:
            return '❌'
    
    def max_chapt(name_path):
        max1 = -np.inf
        max2 = -np.inf
        
        for el in os.listdir(base_path+name_path):
            try:
                chpt = int(chapt_search(el))
                if chpt>max1:
                    max1=chpt
            except:pass
        
        for el in os.listdir(clean_path+name_path+"*"):
            try:
                chpt = int(chapt_search(el))
                if chpt>max2:
                    max2=chpt
            except:pass
        
        return[max1,max2]
            
    if mode == "input_clean_diff":
        df = pd.DataFrame({})
        df["Pas dans 'clean_path'"]=list(s1.difference(s2).difference(s3))
        df.sort_values(by="Pas dans 'clean_path'", inplace=True)
        df.reset_index(inplace=True,drop=True)
        df['In origin'] = df["Pas dans 'clean_path'"].apply(lambda x: x in list(g['Manga_path']))
        df['Statut'] = df["Pas dans 'clean_path'"].apply(lambda x: statut_check(x))
        df.set_index("Pas dans 'clean_path'",drop=True,inplace=True)
        print(df.to_markdown())
        return
    
    elif mode == "check_chapters":
        mg_code = dict(zip(g.Manga_path, g.Manga))
        df = pd.DataFrame({})
        df["Manga_path"] = list(s2.intersection(s1))
        df["Manga"] = df["Manga_path"].apply(lambda x: mg_code[x])
        df.sort_values(by="Manga",inplace=True)
        df.reset_index(inplace=True, drop=True)
        df["input"],df["clean"],df["origin"] = "","",""
        for k in range(len(df)):
            df.loc[k,["input","clean"]] = max_chapt(df.loc[k]["Manga_path"])
            try:
                df.loc[k,"origin"] = int(chapt_origin[df.loc[k]["Manga_path"]])
            except:
                df.loc[k,"origin"] = np.nan
        df.set_index("Manga",drop=True,inplace=True)
        if check_view=="short":
            print(df[df["input"]!=df["clean"]].to_markdown())
        else:
            print(df.to_markdown())
        return

# ========================== #
if __name__ == "__main__":
    print('-------------------------------------------------')
    #try:
    #    size = check_repo_size(git_pth)
    #    print(">(info) Repo. Git plein à",str(round(100*(size/1000/1000),1))+"%\n")
    #except:
    #    pass

    q_menu1 = [
    inquirer.List(name='menu1',
                    message="",
                    choices=['DL',"Edit 'origin.xlsx'",'Check updates','Source HakuNeko',"Open 'origin.xlsx'"]+['↪ Quit'],
                    carousel=True,
                )  
    ]
    a_menu1 = inquirer.prompt(q_menu1, theme=CustomTheme(), raise_keyboard_interrupt=True)["menu1"]

    # ## edit origin.xlsx
    if a_menu1 == "Edit 'origin.xlsx'":
        q_origin = [
            inquirer.List(name='origin',
                    message="",
                    choices=['Existing manga','Add manga']+['↪ Quit'],
                    carousel=True,
                )  
        ]
        a_origin = inquirer.prompt(q_origin, theme=CustomTheme(), raise_keyboard_interrupt=True)["origin"]
        if a_origin == 'Existing manga':
            edit_origin()
            print(term.seashell4(">> Save 'origin.xlsx' et fermer"))
            os.system("open origin.xlsx")
            input(term.seashell4(">> 'Entrer' quand sauvegarde faite: "))
            print('Done ✅')
        elif a_origin == 'Add manga':
            add_origin()
            print(term.seashell4(">> Save 'origin.xlsx' et fermer"))
            os.system("open origin.xlsx")
            input(term.seashell4(">> 'Entrer' quand sauvegarde faite: "))
            print('Done ✅')
        else:
            quit()

    # ## scan download
    elif a_menu1 == "DL":
        main = manga_select()
        print(term.gold3("➜ DL {} - Cover update {} - Scan {} ?".format(*main)))
        conf = [
        inquirer.List(name='conf',
                    message="Confirm",
                    choices=['Yes','No'],
                    carousel=True)
        ]
        conf = inquirer.prompt(conf, theme=CustomTheme(), raise_keyboard_interrupt=True)['conf']
        if conf == "No":
            quit()
        print('-------------------------------------------------')
        if main[1]==True:
            ems_cover_dl(manga = main[0])
        zp = EMS_central(manga=main[0], scan_mode=main[2])
        time.sleep(1)
        try:
            if main[2] == 'TBD update':
                tbd_replace(main[0],zp)
        except:
            pass
    
    elif a_menu1 == "Check updates":
        q_update = [
        inquirer.List(name='update_mode',
                      message="",
                      choices=["Compare input / clean","Check chapt. updates","View"]+['↪ Quit'],
                      carousel=True)
        ]
        a_update = inquirer.prompt(q_update, theme=CustomTheme(), raise_keyboard_interrupt=True)["update_mode"]
        if a_update == "Compare input / clean":
            check_updates(mode="input_clean_diff")
        elif a_update == "Check chapt. updates":
            check_updates(mode="check_chapters")
        elif a_update == "View":
            check_updates(mode="check_chapters", check_view="long")
        else:
            quit()
    
    elif a_menu1 == "Source HakuNeko":
        source_dl()

    elif a_menu1 == "Open 'origin.xlsx'":
        os.system("open origin.xlsx")
    
    else:quit()

    #$
    print(term.steelblue3('\n28ix~ $'))